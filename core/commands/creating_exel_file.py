import openpyxl
from core.bases.model import Product
import asyncio
import os

async def add_sheet_by_id(data: list[Product], id: int):
    try:
        existing_workbook = openpyxl.load_workbook('example.xlsx')
        sheet_name = str(id)
        if sheet_name in existing_workbook.sheetnames:
            return -1
        new_sheet = existing_workbook.create_sheet(title=sheet_name)
        await create_table(new_sheet)
        for product in data:
            await add_product_data(new_sheet, product)

        existing_workbook.save('example.xlsx')
        await copy_sheet(new_sheet.title, f'{id}.xlsx', 'CopiedSheet')
    except Exception as e:
        print(f"Error: {e}")


async def create_table(sheet):
    column_names = ['item_id','name','price','additional_info','link']
    column_data_types = ['str', 'str', 'str', 'str', 'str']
    for col_num, (col_name, col_data_type) in enumerate(zip(column_names, column_data_types), start=1):
        sheet.cell(row=1, column=col_num, value=f"{col_name}: {col_data_type}")
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 40
async def add_product_data(sheet, product):
    sheet.append([product.item_id, product.name, product.price, product.additional_info, product.link])

async def copy_sheet(source_sheet_name, destination_filename, destination_sheet_name):
    source_filename = 'example.xlsx'
    try:
        # Load the source workbook and sheet
        source_workbook = openpyxl.load_workbook(source_filename)
        source_sheet = source_workbook[source_sheet_name]

        # Create a new workbook and copy the sheet to it
        destination_workbook = openpyxl.Workbook()
        destination_sheet = destination_workbook.active
        destination_sheet.title = destination_sheet_name

        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
            destination_sheet.append(row)
        destination_sheet.column_dimensions['A'].width = 10
        destination_sheet.column_dimensions['B'].width = 40
        destination_sheet.column_dimensions['C'].width = 10
        destination_sheet.column_dimensions['D'].width = 40
        destination_workbook.save(destination_filename)
        return 0
    except Exception as e:
        print(f"Error: {e}")
        return -1


async def delete_sheet_by_id(sheet_id: int):
    try:
        filename = 'example.xlsx'
        workbook = openpyxl.load_workbook(filename)

        sheet_name = str(sheet_id)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
            print(f"Sheet with ID {sheet_id} deleted successfully.")
            workbook.save(filename)
        else:
            print(f"Sheet with ID {sheet_id} does not exist.")
    except Exception as e:
        print(f"Error: {e}")

async def delete_file(id: int):
    file_name = f'{int}.xlsx'
    try:
        if os.path.exists(file_name):
            os.remove(file_name)
    except Exception as e:
        print(f"Error: {e}")